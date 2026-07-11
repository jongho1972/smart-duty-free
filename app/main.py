"""Smart Duty Free — 인터넷면세점 3사 가격 비교 웹앱.

브랜드명 + 상품명을 입력하면 신라·롯데·신세계 인터넷면세점에서
동일 상품의 할인율과 가격을 찾아 비교 표로 보여준다.
"""

from __future__ import annotations

import asyncio
import os
import re
from contextlib import asynccontextmanager
from pathlib import Path

from fastapi import Body, FastAPI, Query
from fastapi.responses import JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request

from . import clients

BASE_DIR = Path(__file__).parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


def _load_dotenv() -> None:
    """smart_duty_free/.env 가 있으면 환경변수로 로드(이미 설정된 값은 보존).

    VPS에서는 docker 환경변수가 우선하고, 로컬 개발에서는 .env 를 쓴다.
    """
    env_path = BASE_DIR.parent / ".env"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, val = line.split("=", 1)
        os.environ.setdefault(key.strip(), val.strip())


_load_dotenv()

DEFAULT_RATE = 1500.0  # 환율 미확보 시 KRW 환산 추정용 기본값


def _asset_version() -> str:
    """정적 자산(js/css)의 최신 mtime → 캐시버스팅 버전.

    배포 시 git checkout으로 파일 mtime이 갱신되므로 버전이 바뀌어
    브라우저가 이전 app.js/style.css를 캐시로 재사용하는 문제를 막는다.
    """
    latest = 0
    for p in (BASE_DIR / "static").glob("*"):
        try:
            latest = max(latest, int(p.stat().st_mtime))
        except OSError:
            pass
    return str(latest)


ASSET_VERSION = _asset_version()


def _extract_creds(request: Request) -> dict:
    """요청 헤더에서 면세점 자격증명을 추출한다. 미전송 시 빈 문자열."""
    return {
        "lotte_id": request.headers.get("X-Lotte-Id", ""),
        "lotte_pw": request.headers.get("X-Lotte-Pw", ""),
        "ssg_id":   request.headers.get("X-Ssg-Id",   ""),
        "ssg_pw":   request.headers.get("X-Ssg-Pw",   ""),
    }


@asynccontextmanager
async def lifespan(app: FastAPI):
    yield
    # 종료 시 Playwright 리소스 정리
    try:
        await clients._ssg_browser._reset()
    except Exception:
        pass


app = FastAPI(title="Smart Duty Free", lifespan=lifespan)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")


def choose_keyword(brand: str, product: str) -> str:
    """검색에 가장 효과적인 키워드 선택.

    1순위: 상품명의 모델코드(숫자를 포함한 영숫자 4자+, 예 VVCC25/0RB3447001).
           면세점이 그 모델을 보유하면 정확히 좁혀지고, 미보유면 best_match가 거른다.
    2순위: 모델코드가 없으면 상품명의 가장 긴 토큰(GUSSETED·CRESCENT 등 고유 단어).
           브랜드명보다 고유 단어로 검색해야 원하는 상품이 결과에 직접 들어온다.
           무관 브랜드가 섞여도 best_match가 브랜드 불일치로 거른다.
    3순위: 상품 토큰이 없으면 브랜드명.
    """
    tokens = re.findall(r"[A-Za-z0-9]+", product or "")
    models = [t for t in tokens if len(t) >= 4 and any(c.isdigit() for c in t)]
    if models:
        return max(models, key=len)
    longish = [t for t in tokens if len(t) >= 3]
    if longish:
        return max(longish, key=len)
    return clients._clean_brand(brand) or (product or "").strip() or (brand or "").strip()


def _result_block(shop: str, p) -> dict:
    if p is None:
        return {"shop": shop, "found": False}
    d = p.to_dict()
    d["found"] = True
    return d


async def resolve_lotte(keyword: str, mall: str, lotte_id: str, lotte_pw: str, matcher):
    """롯데 지연 로그인 오케스트레이터.

    비로그인으로 먼저 조회하고, **매칭된 상품의 할인율이 로그인에 가려졌을 때만**
    로그인 후 재조회한다(미보유·할인율 노출·외국몰은 로그인 스킵). 이미 세션이
    캐시돼 있으면 로그인을 유발하지 않고 그 쿠키로 바로 조회한다.

    matcher(products) -> Product|None : 호출측의 매칭 로직(브랜드/토큰/폴백 포함).
    Returns (Product|None, login_required: bool)
      login_required=True : 할인율이 로그인에 막혔으나 자격증명이 없어 정가만 반환.
    """
    # 1) 캐시 세션이 있으면 로그인 유발 없이 그 쿠키로 조회
    cookies, cred_key = clients.peek_lotte_session(lotte_id or None, lotte_pw or None)
    if cookies:
        prods, _ = await asyncio.to_thread(clients.fetch_lotte, keyword, cookies, cred_key, mall)
        return matcher(prods), False

    has_creds = cred_key is not None  # peek는 자격증명이 있을 때만 cred_key를 준다
    # 2) 비로그인 조회
    prods, gated = await asyncio.to_thread(clients.fetch_lotte, keyword, None, None, mall)
    matched = matcher(prods)

    # 3) 미보유 / 게이팅 아님 / 할인율 이미 노출 → 로그인 불필요
    if matched is None or not gated or matched.discount_rate is not None:
        return matched, False

    # 4) 매칭 & 할인율 가려짐 → 자격증명 있으면 로그인 후 재조회
    if not has_creds:
        return matched, True
    cookies, cred_key = await clients.ensure_lotte_login(lotte_id or None, lotte_pw or None)
    if not cookies:
        return matched, True  # 로그인 실패 → 비로그인 결과(정가) 유지
    prods2, _ = await asyncio.to_thread(clients.fetch_lotte, keyword, cookies, cred_key, mall)
    return (matcher(prods2) or matched), False


async def compare(brand: str, product: str,
                  lotte_id: str = "", lotte_pw: str = "",
                  ssg_id: str = "", ssg_pw: str = "",
                  mall: str = "kr") -> dict:
    if mall not in clients.MALLS:
        return {"error": "mall은 kr·cn·en·jp만 지원합니다."}
    ssg_lang = clients.MALLS[mall]["ssg_lang"]
    keyword = choose_keyword(brand, product)
    if not keyword:
        return {"error": "브랜드명 또는 상품명을 입력해 주세요."}

    # 롯데는 지연 로그인: 비로그인 우선 조회 → 매칭 & 할인율 가려짐일 때만 로그인 후 재조회
    lotte_t = resolve_lotte(keyword, mall, lotte_id, lotte_pw,
                            lambda res: clients.best_match(res, brand, product, keyword) if res else None)
    # 신라는 동기(curl_cffi) → 스레드, 신세계는 async(Playwright)
    shilla_t = asyncio.to_thread(clients.fetch_shilla, keyword, mall)
    if ssg_lang:
        ssg_t = clients.fetch_ssg_async(keyword, ssg_id or None, ssg_pw or None, ssg_lang)
    else:
        async def _no_ssg():
            return []
        ssg_t = _no_ssg()
    lotte_res, shilla_r, ssg_r = await asyncio.gather(
        lotte_t, shilla_t, ssg_t, return_exceptions=True)

    def pick(res):
        if isinstance(res, Exception) or not res:
            return None
        return clients.best_match(res, brand, product, keyword)

    if isinstance(lotte_res, Exception):
        lotte, lotte_login_required, lotte_err = None, False, True
    else:
        lotte, lotte_login_required = lotte_res
        lotte_err = False
    shilla = pick(shilla_r)
    ssg = pick(ssg_r)

    # 환율: 롯데가 제공하는 정확한 KRW로 추정(없으면 기본값)
    rate = DEFAULT_RATE
    if lotte and lotte.price_krw and lotte.price_sale:
        rate = lotte.price_krw / lotte.price_sale

    # 신라·신세계 KRW 환산(추정)
    for p in (shilla, ssg):
        if p and p.price_krw is None and p.price_sale is not None:
            p.price_krw = int(round(p.price_sale * rate))
            p.krw_estimated = True  # type: ignore[attr-defined]

    shops = {
        "신라": _result_block("신라", shilla),
        "롯데": _result_block("롯데", lotte),
        "신세계": _result_block("신세계", ssg),
    }
    if not ssg_lang:
        shops["신세계"] = {"shop": "신세계", "found": False, "unsupported": True}
    # krw_estimated 플래그 반영
    for key, p in (("신라", shilla), ("신세계", ssg)):
        if p is not None:
            shops[key]["krw_estimated"] = getattr(p, "krw_estimated", False)
    # 롯데 할인율이 로그인에 막혔으나 자격증명이 없어 정가만 나온 경우 UI 힌트
    shops["롯데"]["login_required"] = lotte_login_required
    # 신세계도 향수·화장품 등은 할인가가 로그인에 가려짐 → 게이팅 시 UI 힌트
    if ssg_lang:
        shops["신세계"]["login_required"] = bool(
            ssg and getattr(ssg, "login_gated", False) and ssg.discount_rate is None)

    errors = {
        "신라": isinstance(shilla_r, Exception),
        "롯데": lotte_err,
        "신세계": bool(ssg_lang) and isinstance(ssg_r, Exception),
    }

    return {
        "query": {"brand": brand, "product": product, "keyword": keyword, "mall": mall},
        "shops": shops,
        "errors": errors,
        "exchange_rate": round(rate, 2),
    }


@app.get("/api/compare")
async def api_compare(
    request: Request,
    brand: str = Query("", description="브랜드명"),
    product: str = Query("", description="상품명/모델"),
    mall: str = Query("kr", description="조회 몰 (kr/cn/en/jp)"),
):
    creds = _extract_creds(request)
    return JSONResponse(await compare(brand.strip(), product.strip(),
                                      mall=mall.strip().lower() or "kr", **creds))


async def compare_by_sku(sku: str,
                          lotte_id: str = "", lotte_pw: str = "",
                          ssg_id: str = "", ssg_pw: str = "",
                          mall: str = "kr") -> dict:
    """SKU 번호 → 신라에서 정확 조회 → REF.NO로 롯데·신세계 검색.

    mall(kr/cn/en/jp) 선택 시 3사 모두 해당 언어몰 기준으로 조회한다.
    SKU·REF.NO는 언어 공통이라 몰이 달라도 같은 상품이 잡힌다.
    신세계는 일문몰이 없어 mall=jp 에서 미지원 처리.
    """
    if mall not in clients.MALLS:
        return {"error": "mall은 kr·cn·en·jp만 지원합니다."}
    ssg_lang = clients.MALLS[mall]["ssg_lang"]

    shilla_product, meta = await asyncio.to_thread(clients.fetch_shilla_by_sku, sku, mall)

    if shilla_product is None:
        mall_label = {"kr": "국문몰", "cn": "중문몰", "en": "영문몰", "jp": "일문몰"}[mall]
        return {"error": f"신라 {mall_label}에서 SKU {sku}를 찾을 수 없습니다."}

    ref_no = meta.get("ref_no", "")
    brand_kr = meta.get("brand_kr", "")
    brand_en = meta.get("brand_en", "")
    product_name = meta.get("product_name", "")
    search_kw = ref_no or choose_keyword(brand_kr, product_name)

    if not search_kw:
        return {"error": "검색 키워드를 추출할 수 없습니다."}

    # 외국몰 결과는 상품명 언어가 달라 토큰 매칭이 약함 → 영문 브랜드 우선 사용,
    # 그래도 못 잡으면 REF.NO 정밀 검색(결과 소수)일 때 최상위 후보를 채택.
    match_brand = brand_kr if mall == "kr" else (brand_en or brand_kr)

    def _match(res):
        if not res:
            return None
        m = clients.best_match(res, match_brand, product_name, search_kw)
        if m is None and mall != "kr" and search_kw == ref_no and len(res) <= 5:
            m = next((p for p in res if not p.soldout), res[0])
        return m

    # 롯데는 지연 로그인(할인율 가려짐 & 매칭될 때만 로그인)
    lotte_t = resolve_lotte(search_kw, mall, lotte_id, lotte_pw, _match)
    if ssg_lang:
        ssg_t = clients.fetch_ssg_async(search_kw, ssg_id or None, ssg_pw or None, ssg_lang)
    else:
        async def _no_ssg():
            return []
        ssg_t = _no_ssg()
    lotte_res, ssg_r = await asyncio.gather(lotte_t, ssg_t, return_exceptions=True)

    def pick(res):
        if isinstance(res, Exception) or not res:
            return None
        return _match(res)

    if isinstance(lotte_res, Exception):
        lotte, lotte_login_required, lotte_err = None, False, True
    else:
        lotte, lotte_login_required = lotte_res
        lotte_err = False
    ssg = pick(ssg_r)

    rate = DEFAULT_RATE
    if lotte and lotte.price_krw and lotte.price_sale:
        rate = lotte.price_krw / lotte.price_sale

    for p in (shilla_product, ssg):
        if p and p.price_krw is None and p.price_sale is not None:
            p.price_krw = int(round(p.price_sale * rate))
            p.krw_estimated = True  # type: ignore[attr-defined]

    shops = {
        "신라": _result_block("신라", shilla_product),
        "롯데": _result_block("롯데", lotte),
        "신세계": _result_block("신세계", ssg),
    }
    if not ssg_lang:
        shops["신세계"] = {"shop": "신세계", "found": False, "unsupported": True}
    for key, p in (("신라", shilla_product), ("신세계", ssg)):
        if p is not None:
            shops[key]["krw_estimated"] = getattr(p, "krw_estimated", False)
    shops["롯데"]["login_required"] = lotte_login_required
    # 신세계도 향수·화장품 등은 할인가가 로그인에 가려짐 → 게이팅 시 UI 힌트
    if ssg_lang:
        shops["신세계"]["login_required"] = bool(
            ssg and getattr(ssg, "login_gated", False) and ssg.discount_rate is None)

    errors = {
        "신라": False,
        "롯데": lotte_err,
        "신세계": bool(ssg_lang) and isinstance(ssg_r, Exception),
    }

    return {
        "query": {
            "sku": sku,
            "ref_no": ref_no,
            "brand": brand_kr,
            "brand_en": brand_en,
            "category": meta.get("category", ""),
            "product": product_name,
            "keyword": search_kw,
            "mall": mall,
        },
        "shops": shops,
        "errors": errors,
        "exchange_rate": round(rate, 2),
    }


@app.get("/api/compare-by-sku")
async def api_compare_by_sku(
    request: Request,
    sku: str = Query("", description="SKU 번호"),
    mall: str = Query("kr", description="조회 몰 (kr/cn/en/jp)"),
):
    creds = _extract_creds(request)
    return JSONResponse(await compare_by_sku(sku.strip(), mall=mall.strip().lower() or "kr", **creds))


EXPORT_HEADERS = [
    "SKU", "국문 브랜드명", "영문 브랜드명", "상품유형", "상품명", "REF.NO",
    "정가(USD)", "신라 할인률", "롯데 할인률", "신세계 할인률",
    "신라 링크", "롯데 링크", "신세계 링크",
]
EXPORT_SHOPS = ["신라", "롯데", "신세계"]


@app.post("/api/export")
async def api_export(payload: dict = Body(default={})):
    """결과를 .xlsx로 생성. 웹 표와 동일하게 할인률과 '가격확인 링크'를
    별도 컬럼으로 구분한다(엑셀은 셀당 하이퍼링크 1개 제약 → 면세점별 링크 컬럼)."""

    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = (payload or {}).get("rows") or []
    wb = Workbook()
    ws = wb.active
    ws.title = "가격비교"
    ws.append(EXPORT_HEADERS)
    head_font = Font(bold=True, color="0B2E5C")
    for c in ws[1]:
        c.font = head_font
    link_font = Font(color="1F6FEB", underline="single")

    for r in rows:
        shops = (r or {}).get("shops") or {}
        ws.append([
            r.get("sku", ""), r.get("brand_kr", ""), r.get("brand_en", ""),
            r.get("category", ""), r.get("product", ""), r.get("ref_no", ""),
            r.get("price_origin", ""), "", "", "", "", "", "",
        ])
        rownum = ws.max_row
        for i, s in enumerate(EXPORT_SHOPS):
            sh = shops.get(s) or {}
            # 할인률(숫자만, 링크 없음)
            rate_cell = ws.cell(row=rownum, column=8 + i)
            rate_cell.value = sh.get("rate") or "—"
            # 가격확인 링크(별도 컬럼) — '바로가기' 텍스트에 하이퍼링크
            link_cell = ws.cell(row=rownum, column=11 + i)
            url = sh.get("url")
            if url:
                link_cell.value = "바로가기"
                link_cell.hyperlink = url
                link_cell.font = link_font
            else:
                link_cell.value = "—"

    for col, width in zip("ABCDEFGHIJKLM", (18, 16, 18, 12, 30, 14, 10, 11, 11, 11, 11, 11, 11)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="price_compare.xlsx"'},
    )


@app.get("/")
async def index(request: Request):
    return templates.TemplateResponse(
        request, "index.html", {"v": ASSET_VERSION})


@app.get("/healthz")
async def healthz():
    return {"ok": True}


@app.post("/api/login-reset")
async def api_login_reset(request: Request):
    """SSG·롯데 로그인 캐시를 초기화해 다음 조회 시 재로그인을 유도."""
    creds = _extract_creds(request)
    clients.invalidate_lotte_login()
    await clients._ssg_browser._reset()
    # 바로 재로그인 시도(SSG)
    ssg_id = creds.get("ssg_id", "")
    ssg_pw = creds.get("ssg_pw", "")
    lotte_id = creds.get("lotte_id") or os.getenv("LOTTE_ID", "")
    lotte_pw = creds.get("lotte_pw") or os.getenv("LOTTE_PW", "")
    ssg_ok = False
    lotte_ok = False
    if ssg_id and ssg_pw:
        try:
            await clients._ssg_browser._ensure()
            await clients._ssg_browser._ensure_login(ssg_id, ssg_pw)
            ssg_ok = clients._ssg_browser._logged_in
        except Exception:
            pass
    lotte_dbg: dict = {}
    try:
        jar, lotte_dbg = await clients._do_lotte_login(
            lotte_id or os.getenv("LOTTE_ID", ""),
            lotte_pw or os.getenv("LOTTE_PW", ""),
        )
        lotte_ok = jar is not None
    except Exception as e:
        lotte_dbg["exception"] = str(e)
    return JSONResponse({"ssg_login": ssg_ok, "lotte_login": lotte_ok, "lotte_dbg": lotte_dbg})


_ROBOTS_TXT = (
    "User-agent: *\n"
    "Allow: /\n"
    "Disallow: /api/\n\n"
    "Sitemap: https://dfprice.jhawk.kr/sitemap.xml\n"
)

_SITEMAP_XML = (
    '<?xml version="1.0" encoding="UTF-8"?>\n'
    '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
    "  <url>\n"
    "    <loc>https://dfprice.jhawk.kr/</loc>\n"
    "    <changefreq>weekly</changefreq>\n"
    "    <priority>1.0</priority>\n"
    "  </url>\n"
    "</urlset>\n"
)


@app.get("/robots.txt", include_in_schema=False)
async def robots_txt():
    return Response(_ROBOTS_TXT, media_type="text/plain")


@app.get("/sitemap.xml", include_in_schema=False)
async def sitemap_xml():
    return Response(_SITEMAP_XML, media_type="application/xml")
